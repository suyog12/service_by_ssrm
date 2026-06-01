import datetime
import os
from collections import defaultdict

_results = []


def pytest_runtest_logreport(report):
    if report.when == "call":
        _results.append({
            "node_id":  report.nodeid,
            "outcome":  report.outcome.upper(),
            "duration": round(report.duration, 3),
            "error":    str(report.longrepr) if report.failed else ""
        })


def pytest_sessionfinish(session, exitstatus):
    if not _results:
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        HDR_BG  = "1F2937"; HDR_FG  = "FFFFFF"
        MOD_BG  = "374151"
        PASS_BG = "D1FAE5"; PASS_FG = "065F46"
        FAIL_BG = "FEE2E2"; FAIL_FG = "991B1B"
        ALT     = "F9FAFB"; WHITE   = "FFFFFF"

        def fill(h):
            return PatternFill("solid", start_color=h, fgColor=h)

        def fnt(color="111827", bold=False, size=9):
            return Font(color=color, bold=bold, size=size, name="Arial")

        def bdr():
            s = Side(style="thin", color="E5E7EB")
            return Border(left=s, right=s, top=s, bottom=s)

        def ctr():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def lft():
            return Alignment(horizontal="left", vertical="center", wrap_text=True)

        total     = len(_results)
        passed    = sum(1 for r in _results if r["outcome"] == "PASSED")
        failed    = sum(1 for r in _results if r["outcome"] == "FAILED")
        run_at    = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        pass_rate = round(passed / total * 100, 1) if total else 0

        # Sheet 1: All Results 
        ws = wb.active
        ws.title = "Results"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Service by SSRM — Test Results  |  {run_at}"
        ws["A1"].font = Font(name="Arial", size=13, bold=True, color=HDR_FG)
        ws["A1"].fill = fill(HDR_BG)
        ws["A1"].alignment = ctr()
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"] = (
            f"Total: {total}   |   Passed: {passed}   |   "
            f"Failed: {failed}   |   Pass rate: {pass_rate}%"
        )
        ws["A2"].font = Font(
            name="Arial", size=10,
            color=PASS_FG if failed == 0 else FAIL_FG
        )
        ws["A2"].fill = fill(PASS_BG if failed == 0 else FAIL_BG)
        ws["A2"].alignment = ctr()
        ws.row_dimensions[2].height = 18

        headers = ["#", "Module", "Test Class", "Test Name",
                   "Result", "Duration (s)", "Error"]
        widths  = [5, 22, 28, 44, 10, 14, 55]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = fnt(HDR_FG, bold=True, size=10)
            c.fill = fill(MOD_BG)
            c.alignment = ctr()
            c.border = bdr()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[3].height = 18

        for idx, r in enumerate(_results, 1):
            parts      = r["node_id"].replace("\\", "/").split("::")
            file_part  = parts[0]
            test_class = parts[1] if len(parts) > 1 else ""
            test_name  = parts[2] if len(parts) > 2 else parts[-1]
            segments   = file_part.split("/")
            module     = (segments[1].replace("_", " ").title()
                          if len(segments) > 1 else file_part)

            rb = ALT if idx % 2 == 0 else WHITE
            row_data = [idx, module, test_class, test_name,
                        r["outcome"], r["duration"],
                        r["error"][:500] if r["error"] else ""]

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=idx + 3, column=col, value=val)
                c.border = bdr()
                c.alignment = lft()
                c.font = fnt(size=9)
                c.fill = fill(rb)

            rc = ws.cell(row=idx + 3, column=5)
            rc.fill = fill(PASS_BG if r["outcome"] == "PASSED" else FAIL_BG)
            rc.font = fnt(PASS_FG if r["outcome"] == "PASSED" else FAIL_FG,
                          bold=True, size=9)
            rc.alignment = ctr()
            ws.row_dimensions[idx + 3].height = 16

        # Sheet 2: Module Summary 
        ws2 = wb.create_sheet("Module Summary")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:E1")
        ws2["A1"] = "Results by Module"
        ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=HDR_FG)
        ws2["A1"].fill = fill(HDR_BG)
        ws2["A1"].alignment = ctr()
        ws2.row_dimensions[1].height = 28

        for i, (h, w) in enumerate(zip(
            ["Module", "Total", "Passed", "Failed", "Pass Rate"],
            [28, 10, 10, 10, 14]
        ), 1):
            c = ws2.cell(row=2, column=i, value=h)
            c.font = fnt(HDR_FG, bold=True, size=10)
            c.fill = fill(MOD_BG)
            c.alignment = ctr()
            c.border = bdr()
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.row_dimensions[2].height = 18

        mod_stats = defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0})
        for r in _results:
            segs = r["node_id"].replace("\\", "/").split("/")
            mod  = segs[1].replace("_", " ").title() if len(segs) > 1 else "Other"
            mod_stats[mod]["total"] += 1
            if r["outcome"] == "PASSED":
                mod_stats[mod]["passed"] += 1
            else:
                mod_stats[mod]["failed"] += 1

        for row_n, (mod, s) in enumerate(mod_stats.items(), 3):
            rb   = ALT if row_n % 2 == 0 else WHITE
            rate = round(s["passed"] / s["total"] * 100, 1) if s["total"] else 0

            for col in range(1, 6):
                ws2.cell(row=row_n, column=col).fill = fill(rb)
                ws2.cell(row=row_n, column=col).border = bdr()

            ws2.cell(row=row_n, column=1).value = mod
            ws2.cell(row=row_n, column=1).font = fnt(size=10)
            ws2.cell(row=row_n, column=1).alignment = lft()

            ws2.cell(row=row_n, column=2).value = s["total"]
            ws2.cell(row=row_n, column=2).font = fnt(bold=True)
            ws2.cell(row=row_n, column=2).alignment = ctr()

            pc = ws2.cell(row=row_n, column=3, value=s["passed"])
            pc.fill = fill(PASS_BG); pc.font = fnt(PASS_FG, bold=True)
            pc.alignment = ctr(); pc.border = bdr()

            fc = ws2.cell(row=row_n, column=4, value=s["failed"])
            fc.fill = fill(FAIL_BG if s["failed"] > 0 else PASS_BG)
            fc.font = fnt(FAIL_FG if s["failed"] > 0 else PASS_FG, bold=True)
            fc.alignment = ctr(); fc.border = bdr()

            if rate == 100:
                rate_bg, rate_fg = PASS_BG, PASS_FG
            elif rate < 80:
                rate_bg, rate_fg = FAIL_BG, FAIL_FG
            else:
                rate_bg, rate_fg = "FEF3C7", "92400E"

            rc2 = ws2.cell(row=row_n, column=5, value=f"{rate}%")
            rc2.fill = fill(rate_bg); rc2.font = fnt(rate_fg, bold=True)
            rc2.alignment = ctr(); rc2.border = bdr()
            ws2.row_dimensions[row_n].height = 18

        last = len(mod_stats) + 3
        for col in range(1, 6):
            ws2.cell(row=last, column=col).fill = fill(HDR_BG)
            ws2.cell(row=last, column=col).border = bdr()
        for i, v in enumerate(["TOTAL", total, passed, failed, f"{pass_rate}%"], 1):
            c = ws2.cell(row=last, column=i, value=v)
            c.font = Font(name="Arial", color=HDR_FG, bold=True, size=10)
            c.alignment = ctr() if i > 1 else lft()
        ws2.row_dimensions[last].height = 22

        timestamp   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.path.dirname(__file__),
            f"test_results_{timestamp}.xlsx"
        )
        wb.save(output_path)
        print(f"\n  Excel report saved: {output_path}\n")

    except Exception as e:
        print(f"\n  Could not write Excel report: {e}\n")